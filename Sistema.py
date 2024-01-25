import cv2
from pyzbar.pyzbar import decode
import numpy as np
from datetime import datetime
import openpyxl as xl


cap = cv2.VideoCapture(0)

mañana = []
tarde = []
noche = []

def infhora():
    inf = datetime.now()
    fecha = inf.strftime('%Y:%m:%d')
    hora = inf.strftime('%H:%M:%S')
    return hora, fecha

while True:
    # Leemos los frames
    ret, frame = cap.read()

    # INTERFAZ
    # Texto
    cv2.putText(frame, 'Locate the QR code', (160, 80), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
    # Ubicanos el rectangulo en las zonas
    cv2.rectangle(frame, (170, 100), (470, 400), (0, 255, 0), 2)

    # Extraemos hora y fecha
    hora, fecha = infhora()
    diasem = datetime.today().weekday()
    print(diasem)

    # Año | Mes | Dia
    a, me, d = fecha[0:4], fecha[5:7], fecha[8:10]
    # Hora | Minuto | Segundo
    h, m, s = int(hora[0:2]), int(hora[3:5]), int(hora[6:8])

    # Creamos archivo
    nomar = f"{a}-{me}-{d}"
    texth = f"{h}:{m}:{s}"
    print(texth)
    wb = xl.Workbook()

    # LEEMOS LOS CODIGOS QR
    for codes in decode(frame):

        # INFORMACION
        # Decodidficamos
        info = codes.data.decode('utf-8')
        # Tipo de persona LETRA
        tipo = int(info[0:2])
        letr = chr(tipo)

        # Numero
        num = info[2:]
        
        nombre = info.split('-')[1].strip()

        # Extraemos coordenadas
        pts = np.array([codes.polygon], np.int32)
        xi, yi = codes.rect.left, codes.rect.top

        # Redimensionamos
        pts = pts.reshape((-1, 1, 2))
        # id completo
        codigo = letr + num

        # DIAS DE LA SEMANA
        # SEMANA
        if 6 >= diasem >= 0:

            # DIVIDINOS LAS HORAS DEL DIA
            # MAÑANA
            if 12 >= h >= 7:
                cv2.polylines(frame, [pts], True, (255, 255, 0), 5)

                # Guardamos ID, fecha y hora
                if codigo not in mañana:
                    mañana.append(codigo)

                    hojam = wb.create_sheet("Mañana")

                    # Añadimos encabezados si la hoja está vacía
                    if hojam.max_row == 1:
                        hojam.append(["ID", "Nombre", "Fecha", "Hora"])  # Encabezados

                    codigo_despues_del_primer_digito = info.split('-')[0][2:].strip()

                    hojam.append([codigo_despues_del_primer_digito, nombre, fecha, texth])  # Agregamos ID, fecha y hora
                    wb.save(f"{nomar}.xlsx")

                    cv2.putText(frame, f"{letr}0{num}", (xi - 15, yi - 15), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 55, 0), 2)
                    print("El usuario es accionista de la empresa \nNúmero de Identificación:", codigo, "Fecha de registro:", fecha, "Hora de registro:", texth)

                elif codigo in mañana:
                    cv2.putText(frame, f"El ID {codigo}", (xi - 65, yi - 45), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 0, 0), 2)
                    cv2.putText(frame, "Fue registrado", (xi - 65, yi - 15), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 0, 0), 2)
                    print(mañana)


            # TARDE
            elif 18 >= h >= 12:
                cv2.polylines(frame, [pts], True, (255, 255, 0), 5)

                # Guardamos ID, fecha y hora
                if codigo not in tarde:
                    tarde.append(codigo)

                    hojat = wb.create_sheet("Tarde")

                    # Añadimos encabezados si la hoja está vacía
                    if hojat.max_row == 1:
                        hojat.append(["ID", "Nombre", "Fecha", "Hora"])  # Encabezados
                    
                    codigo_despues_del_primer_digito = info.split('-')[0][2:].strip()

                    hojat.append([codigo_despues_del_primer_digito, nombre, fecha, texth])
                    wb.save(f"{nomar}.xlsx")

                    cv2.putText(frame, f"{letr}0{num}", (xi - 15, yi - 15), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 55, 0), 2)
                    print("El usuario es accionista de la empresa \nNúmero de Identificación:", codigo, "Fecha de registro:", fecha, "Hora de registro:", texth)

                elif codigo in tarde:
                    cv2.putText(frame, f"El ID {codigo}", (xi - 65, yi - 45), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
                    cv2.putText(frame, "Fue registrado", (xi - 65, yi - 15), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
            
            # NOCHE
            if 23 >= h >= 18:
                cv2.polylines(frame, [pts], True, (255, 255, 0), 5)

                # Guardamos ID, fecha y hora
                if codigo not in noche:
                    noche.append(codigo)

                    hojan = wb.create_sheet("Noche")

                    # Añadimos encabezados si la hoja está vacía
                    if hojan.max_row == 1:
                        hojan.append(["ID", "Nombre", "Fecha", "Hora"])  # Encabezados
                        
                    codigo_despues_del_primer_digito = info.split('-')[0][2:].strip()

                    hojan.append([codigo_despues_del_primer_digito, nombre, fecha, texth])  # Agregamos ID, fecha y hora
                    wb.save(f"{nomar}.xlsx")

                    cv2.putText(frame, f"{letr}0{num}", (xi - 15, yi - 15), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 55, 0), 2)
                    print("El usuario es accionista de la empresa \nNúmero de Identificación:", codigo, "Fecha de registro:", fecha, "Hora de registro:", texth)

                elif codigo in noche:
                    cv2.putText(frame, f"El ID {codigo}", (xi - 65, yi - 45), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)
                    cv2.putText(frame, "Fue registrado", (xi - 65, yi - 15), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)
                    print(noche)


    # Mostramos FPS
    cv2.imshow("SISTEMA CONTROL DE ACCESO", frame)

    # Leemos teclado
    t = cv2.waitKey(5)
    if t == 27:
        break

cv2.destroyAllWindows()
cap.release()
