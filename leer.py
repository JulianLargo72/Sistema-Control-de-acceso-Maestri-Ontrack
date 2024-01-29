import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.application import MIMEApplication
import os
from flask import Flask, render_template, url_for, flash, request, redirect, Response
import cv2
from pyzbar.pyzbar import decode
import numpy as np
from datetime import datetime
import openpyxl as xl
import time
import pyqrcode

app = Flask(__name__)
app.secret_key = 'maestri'

mañana = []
tiempo_ultima_registro = {}

# Diccionario para mapear el primer carácter al área correspondiente
areas = {
    'A': 'Administrativa',
    'G': 'Gerencial',
    'C': 'Comercial',
    'O': 'Operaciones'
}

def obtener_area(codigo):
    primer_caracter = codigo[0]
    return areas.get(primer_caracter, 'Desconocida')

def infhora():
    inf = datetime.now()
    fecha = inf.strftime('%Y-%m-%d')
    hora = inf.strftime('%H:%M:%S')
    return hora, fecha

# Especificar el nombre de la nueva carpeta para los registros de Excel
nombre_carpeta_registros = 'registros_excel'

# Crear la carpeta si no existe
if not os.path.exists(nombre_carpeta_registros):
    os.makedirs(nombre_carpeta_registros)

# Construir la ruta completa al archivo Excel en la nueva carpeta
hora, fecha_actual = infhora()
ruta_archivo_excel = os.path.join(nombre_carpeta_registros, f"{fecha_actual}.xlsx")

try:
    wb = xl.load_workbook(ruta_archivo_excel)
    hojam = wb["Actual"]
except FileNotFoundError:
    # Si el archivo no existe, crear uno nuevo en la nueva carpeta
    wb = xl.Workbook()
    hojam = wb.create_sheet("Actual")
    hojam.append(["Identificacion", "Nombre", "Area", "Fecha", "Hora Escaneo", "Hora Entrada", "Hora Salida"])  # Encabezados
    # Guardar en la nueva carpeta
    wb.save(ruta_archivo_excel)

def generate_frames():
    cap = cv2.VideoCapture(1)
    while True:
        ret, frame = cap.read()
        
        # INTERFAZ
        # Texto
        cv2.putText(frame, 'Ubica el QR code', (160, 80), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
        # Ubicanos el rectangulo en las zonas
        cv2.rectangle(frame, (170, 100), (470, 400), (0, 255, 0), 2)

        hora, fecha = infhora()
        diasem = datetime.today().weekday()

        a, me, d = fecha[0:4], fecha[5:7], fecha[8:10]
        h, m, s = int(hora[0:2]), int(hora[3:5]), int(hora[6:8])

        # Obtener la hora actual
        hora_actual = datetime.now().time()

        # Formatear la hora, minutos y segundos con dos dígitos
        texth = hora_actual.strftime('%H:%M:%S')
        print(texth)

        for codes in decode(frame):
            info = codes.data.decode('utf-8')
            tipo = int(info[0:2])
            letr = chr(tipo)
            num = info[2:]
            nombre = info.split('-')[1].strip()
            pts = np.array([codes.polygon], np.int32)
            xi, yi = codes.rect.left, codes.rect.top
            pts = pts.reshape((-1, 1, 2))
            codigo = letr + num
            
            # Imprimir el valor del código QR en la consola
            print("Valor del código QR:", codigo)

            if 6 >= diasem >= 0:
                cv2.polylines(frame, [pts], True, (255, 255, 0), 5)

                if codigo not in mañana or (codigo in mañana and time.time() - tiempo_ultima_registro.get(codigo, 0) > 60):
                    mañana.append(codigo)
                    tiempo_ultima_registro[codigo] = time.time()

                    # Obtener el área correspondiente al primer carácter del código
                    area = obtener_area(codigo)

                    codigo_despues_del_primer_digito = info.split('-')[0][2:].strip()

                    # Buscar la hora de entrada en el archivo de Excel
                    hora_entrada = None
                    for row in hojam.iter_rows(min_row=2, max_col=6, max_row=hojam.max_row):
                        if row[0].value == codigo_despues_del_primer_digito:
                            # Verificar que la tupla tenga suficientes elementos antes de intentar acceder al índice 5
                            if len(row) > 5:
                                hora_entrada = row[5].value
                            break

                    if hora_entrada is None:
                        # Si no se encuentra la hora de entrada, usar la hora actual
                        hora_entrada = time.strftime("%H:%M:%S", time.localtime())

                    # Agregar la fila al archivo de Excel con código, nombre, área, fecha, hora de entrada y hora de salida
                    hojam.append([codigo_despues_del_primer_digito, nombre, area, fecha, texth, hora_entrada, time.strftime("%H:%M:%S")])

                    # Guardar los cambios en el archivo de Excel
                    wb.save(ruta_archivo_excel)


                    cv2.putText(frame, f"{letr}0{num}", (xi - 15, yi - 15), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 55, 0), 2)
                    print("El usuario es accionista de la empresa \nNúmero de Identificación:", codigo, "Fecha de registro:", fecha, "Hora de registro:", texth)

                elif codigo in mañana:
                    cv2.putText(frame, f"El ID {codigo}", (xi - 65, yi - 45), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 0, 0), 2)
                    cv2.putText(frame, "Registro Exitoso", (xi - 65, yi - 15), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 0, 0), 2)
                    print(mañana)

        ret, jpeg = cv2.imencode('.jpg', frame)
        frame_bytes = jpeg.tobytes()
        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + frame_bytes + b'\r\n\r\n')

@app.route('/leer')
def mostrar_leer():
    return render_template('leer.html')

@app.route('/')
def inicio():
    return render_template('index.html')


#Función para generar el código QR y devolver la ruta y la información completa
def generar_qr(prefijo, identificacion, nombre):
    if not prefijo.isalpha():
        raise ValueError("El prefijo debe contener solo letras.")
    
    ascii_prefijo = str(ord(prefijo))  # Convertir letra a código ASCII
    id_completo = ascii_prefijo + str(identificacion)
    info_completa = f"{id_completo} - {nombre}"
    qr = pyqrcode.create(info_completa, error='L')
    qr_path = f'static/qr_images/{prefijo}{identificacion}.png'
    qr.png(qr_path, scale=6)
    return qr_path, info_completa  # Devolver la ruta y la información completa

# Función para enviar el correo electrónico
def enviar_correo(destinatario, asunto, cuerpo, adjunto_path):
    remitente = "sistemasmaestri@gmail.com"
    password = "ugtdrarcnbgsxnbu"

        # Configuración del servidor SMTP de Gmail
    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.starttls()
    server.login(remitente, password)

    # Construir el mensaje del correo electrónico
    mensaje = MIMEMultipart()
    mensaje['From'] = remitente
    mensaje['To'] = destinatario
    mensaje['Subject'] = asunto

    mensaje.attach(MIMEText(cuerpo, 'plain'))

    # Adjuntar la imagen del código QR al mensaje
    with open(adjunto_path, 'rb') as archivo_adjunto:
        adjunto = MIMEImage(archivo_adjunto.read(), name='qr.png')
        mensaje.attach(adjunto)

    # Enviar el correo electrónico
    server.sendmail(remitente, destinatario, mensaje.as_string())
    server.quit()

@app.route('/generar', methods=['GET', 'POST'])
def generar():
    if request.method == 'POST':
        try:
            prefijo = request.form['prefijo']
            if not prefijo.isalpha():
                raise ValueError("El prefijo debe contener solo letras.")
            
            identificacion = int(request.form['identificacion'])
            nombre = request.form['nombre']
            qr_path, info_completa = generar_qr(prefijo, identificacion, nombre)

            # Verificar si el campo de correo está presente en el formulario
            destinatario = request.form.get('correo', '')  # Si no se proporciona, el valor será una cadena vacía

            # Enviar el correo electrónico solo si se proporciona el correo
            if destinatario:
                enviar_correo(destinatario, "Codigo QR control de acceso Maestri Ontrack", f" Cordial saludo:\n\nTe informamos que el QR generado con la siguiente informacion ha sido registrado con exito!\n\nNombre: {nombre}\nIdentificación: {identificacion}\n\nA continuacion adjuntamos el codigo", qr_path)

            mensaje = f"QR generado para la identificación {identificacion} y el nombre {nombre}."
            if destinatario:
                mensaje += f" Se ha enviado al correo {destinatario}."
            return render_template('generar.html', mensaje=mensaje, qr_path=qr_path)
        except ValueError as e:
            mensaje = str(e)
            return render_template('generar.html', mensaje=mensaje)
    return render_template('generar.html')

@app.route('/video_feed')
def video_feed():
    return Response(generate_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')

def enviar_correo_excel(destinatario, asunto, cuerpo, adjunto_path):
    remitente = "sistemasmaestri@gmail.com"
    password = "ugtdrarcnbgsxnbu"

    # Configuración del servidor SMTP de Gmail
    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.starttls()
    server.login(remitente, password)

    # Carga el archivo adjunto en memoria
    archivo_adjunto_memoria = open(adjunto_path, 'rb').read()

    # Crea el mensaje de correo electrónico
    mensaje = MIMEMultipart()
    mensaje['From'] = remitente
    mensaje['To'] = destinatario
    mensaje['Subject'] = asunto

    # Adjuntar el archivo Excel al mensaje
    adjunto = MIMEApplication(archivo_adjunto_memoria, _subtype="xlsx")
    adjunto.add_header('Content-Disposition', f'attachment; filename={os.path.basename(adjunto_path)}')
    mensaje.attach(adjunto)

    # Enviar el correo electrónico
    server.sendmail(remitente, destinatario, mensaje.as_string())
    server.quit()

@app.route('/registros', methods=['GET'])
def mostrar_registros():
    fecha_filtro = request.args.get('fecha', datetime.today().strftime('%Y-%m-%d'))
    carpeta_registros = 'registros_excel'
    
    registros = []

    for archivo_excel in os.listdir(carpeta_registros):
        ruta_archivo_excel = os.path.join(carpeta_registros, archivo_excel)

        try:
            wb = xl.load_workbook(ruta_archivo_excel)
            hoja_actual = wb['Actual']

            for row in hoja_actual.iter_rows(min_row=2, values_only=True):
                registro = {
                    'Identificacion': row[0],
                    'Nombre': row[1],
                    'Area': row[2],
                    'Fecha': row[3],
                    'Hora': row[4],
                    'Entrada': row[5],
                    'Salida': row[6]
                }

                # Filtra los registros por fecha si se proporciona una fecha de filtro
                if fecha_filtro:
                    if registro['Fecha'] == fecha_filtro:
                        registros.append(registro)
                else:
                    registros.append(registro)

        except Exception as e:
            print(f"Error al procesar el archivo {archivo_excel}: {e}")

    return render_template('registros.html', registros=registros, fecha_actual=datetime.today().strftime('%Y-%m-%d'))



@app.route('/enviar_excel', methods=['GET', 'POST'])
def enviar_excel():
    carpeta_registros = 'registros_excel'
    fecha_filtro = datetime.today().strftime('%Y-%m-%d')
    registros = []  # Inicializa la variable registros aquí

    if request.method == 'GET':
        fecha_filtro = request.args.get('fecha', fecha_filtro)

        try:
            archivos_excel = [archivo for archivo in os.listdir(carpeta_registros) if archivo.endswith('.xlsx')]
        except FileNotFoundError:
            flash('No se encontraron archivos de registro.')
            return render_template('enviar_correo.html', registros=registros, fecha_actual=datetime.today().strftime('%Y-%m-%d'))

        for archivo_excel in archivos_excel:
            ruta_archivo_excel = os.path.join(carpeta_registros, archivo_excel)

            try:
                wb = xl.load_workbook(ruta_archivo_excel)
                hoja_actual = wb['Actual']

                for row in hoja_actual.iter_rows(min_row=2, values_only=True):
                    registro = {
                        'Identificacion': row[0],
                        'Nombre': row[1],
                        'Area': row[2],
                        'Fecha': row[3],
                        'Hora': row[4],
                        'Entrada': row[5],
                        'Salida': row[6]
                    }

                    if fecha_filtro:
                        if registro['Fecha'] == fecha_filtro:
                            registros.append(registro)
                    else:
                        registros.append(registro)

            except Exception as e:
                print(f"Error al procesar el archivo {archivo_excel}: {e}")

        return render_template('enviar_correo.html', registros=registros, fecha_actual=datetime.today().strftime('%Y-%m-%d'))

    elif request.method == 'POST':
        correo_destinatario = request.form.get('correo_destinatario')
        fecha_filtro = request.form.get('fecha', fecha_filtro)

        archivo_a_enviar = os.path.join(carpeta_registros, f'{fecha_filtro}.xlsx')

        try:
            enviar_correo_excel(correo_destinatario, 'Archivo Control de Acceso', f"Cordial saludo,\n\nA continuacion adjuntamos el archivo de control de acceso correspondiente al registro de la fecha {fecha_filtro}", archivo_a_enviar)
            flash('Correo enviado exitosamente.', 'success')  # Agrega la alerta de éxito
        except FileNotFoundError:
            flash('El archivo de registro no se encontró.')
            return redirect(url_for('enviar_excel'))

        # Devuelve la plantilla con la fecha actual y la alerta de éxito
        return render_template('enviar_correo.html', registros=[], fecha_actual=datetime.today().strftime('%Y-%m-%d'))

if __name__ == '__main__':
    app.run(host='192.168.0.44', port=5000, debug=True)